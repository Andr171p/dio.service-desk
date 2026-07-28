from __future__ import annotations

from typing import TYPE_CHECKING, Any

from copy import copy
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .common import (
    EXCEL_TEMPLATE_PATH,
    build_report_context,
    ensure_template_exists,
    render_template_string,
)

if TYPE_CHECKING:
    from ..services.stage_report import ProjectStagesReport

ROW_TEMPLATE_MARKER = "{{ row.number }}"


def export_project_stages_to_excel(report: ProjectStagesReport) -> bytes:
    """
    Сформировать Excel-отчёт по этапам проекта из шаблона.
    """

    ensure_template_exists(EXCEL_TEMPLATE_PATH)
    workbook = load_workbook(EXCEL_TEMPLATE_PATH)
    worksheet = workbook.active
    context = build_report_context(report)

    template_row_index = _find_template_row(worksheet)
    _render_report_metadata(worksheet, context, excluded_row=template_row_index)
    _render_stage_rows(
        worksheet=worksheet,
        template_row_index=template_row_index,
        rows=context["rows"],
    )
    _set_project_hyperlink(worksheet, report.project_url)
    _update_print_and_filter_ranges(worksheet, template_row_index, len(report.rows))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _find_template_row(worksheet: Worksheet) -> int:
    for row in worksheet.iter_rows():
        if any(cell.value == ROW_TEMPLATE_MARKER for cell in row):
            return row[0].row

    raise ValueError(
        f"Row marker {ROW_TEMPLATE_MARKER!r} not found in {EXCEL_TEMPLATE_PATH}"
    )


def _render_report_metadata(
    worksheet: Worksheet,
    context: dict[str, Any],
    excluded_row: int,
) -> None:
    for row in worksheet.iter_rows():
        if row[0].row == excluded_row:
            continue

        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                cell.value = render_template_string(cell.value, context)


def _render_stage_rows(
    worksheet: Worksheet,
    template_row_index: int,
    rows: list[dict[str, Any]],
) -> None:
    template_cells = list(worksheet[template_row_index])
    template_height = worksheet.row_dimensions[template_row_index].height

    if not rows:
        worksheet.delete_rows(template_row_index)
        return

    if len(rows) > 1:
        worksheet.insert_rows(
            template_row_index + 1,
            amount=len(rows) - 1,
        )

    for row_offset, row_context in reversed(list(enumerate(rows))):
        target_row_index = template_row_index + row_offset

        worksheet.row_dimensions[target_row_index].height = template_height
        for template_cell in template_cells:
            target_cell = worksheet.cell(
                row=target_row_index,
                column=template_cell.column,
            )
            _copy_cell_style(template_cell, target_cell)
            if isinstance(template_cell.value, str):
                target_cell.value = render_template_string(
                    template_cell.value,
                    {"row": row_context},
                )
            else:
                target_cell.value = template_cell.value


def _copy_cell_style(source: Cell, target: Cell) -> None:
    if source.has_style:
        target._style = copy(source._style)  # noqa: SLF001

    target.number_format = source.number_format
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)


def _set_project_hyperlink(worksheet: Worksheet, project_url: str) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and project_url in cell.value:
                cell.hyperlink = project_url
                return


def _update_print_and_filter_ranges(
    worksheet: Worksheet,
    template_row_index: int,
    rows_count: int,
) -> None:
    header_row_index = template_row_index - 1
    last_row_index = header_row_index + rows_count
    last_column_letter = worksheet.cell(
        row=header_row_index,
        column=worksheet.max_column,
    ).column_letter

    worksheet.auto_filter.ref = (
        f"A{header_row_index}:{last_column_letter}{max(header_row_index, last_row_index)}"
    )
    worksheet.print_area = (
        f"A1:{last_column_letter}{max(header_row_index, last_row_index)}"
    )
